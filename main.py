from config import token_b, mng_id, s_k_id, path_dir
from telebot import types
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from datetime import date
import openpyxl as opx
from random import randint

import pprint
import re
import os
import threading
import schedule
import telebot
import traceback
import time
import os.path
import sqlite3

# main data
try:
    os.chdir(path_dir)
except:
    pass
bot = telebot.TeleBot(token_b, threaded=False)
message_dict = {}
user_dict = {}
path_to_n_cont = 'n_count.txt'
path_n_global = 'n_count_global.txt'
driver_file_name = "driver_list"
notif = 'notif.txt'
change = ''
time_change = 0


# keyboards
# cancel button


def cancel():
    cancel = types.InlineKeyboardButton(text='Отмена', callback_data='cancel')
    return cancel


# к
# клавиатура с админскими функциями
def kb_cmd_admin():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    refresh = types.KeyboardButton('Обновить список водителей')
    med_meh = types.KeyboardButton('Обновить список работников')
    replace = types.KeyboardButton('Замена сотрудника')
    write = types.KeyboardButton('Рассылка')
    back = types.KeyboardButton('Вернуться')
    markup.add(refresh, write)
    markup.add(med_meh, replace)
    markup.add(back)
    return markup


# keyboard reg
def reg_keyboard():
    inline_keyboard_reg = types.InlineKeyboardMarkup()
    reg_1 = types.InlineKeyboardButton(text='Обновить', callback_data='reg_change')
    reg_2 = types.InlineKeyboardButton(text='Удалить', callback_data='reg_del')
    inline_keyboard_reg.add(reg_1, reg_2)
    inline_keyboard_reg.add(cancel())
    return inline_keyboard_reg


# клавиатура достающая список медиков и механиков для замены
def worker_kb():
    inline_keyboard_worker = types.InlineKeyboardMarkup()
    wb = load_workbook(f'ТабельМЕДИК-МЕХАНИК.xlsx')
    ws = wb.active
    for cell in ws.iter_rows(min_col=1, max_col=1):
        if cell[0].value == None:
            continue
        btn_name = types.InlineKeyboardButton(text=f'{cell[0].value}', callback_data=f'worker_{cell[0].value}')
        inline_keyboard_worker.add(btn_name)
    return inline_keyboard_worker

# клавиатура достающая список медиков и механиков для замены
def worker_change_kb():
    inline_keyboard_changer = types.InlineKeyboardMarkup()
    wb = load_workbook(f'ТабельМЕДИК-МЕХАНИК.xlsx')
    ws = wb.active
    btn_noone = types.InlineKeyboardButton(text=f'Никого', callback_data=f'chw_noone')
    for cell in ws.iter_rows(min_col=1, max_col=1):
        if cell[0].value == None:
            continue
    # chw - change worker
        btn_name = types.InlineKeyboardButton(text=f'{cell[0].value}', callback_data=f'chw_{cell[0].value}')
        inline_keyboard_changer.add(btn_name)
    inline_keyboard_changer.add(btn_noone)
    return inline_keyboard_changer

def phone_btn():
    kb = types.ReplyKeyboardMarkup(row_width=1)
    btn_send_phone = types.KeyboardButton(text='Отправить номер', request_contact=True)
    kb.add(btn_send_phone)
    return kb


# клава для выбора времени осмотра медиком
def time_med_kb(time_set=time.time(), time_type='me_', person=None, first=False):
    global change
    if first:
        change = person
    time_btn = time.strftime("%H:%M", time.localtime(time_set))
    inline_keyboard_time = types.InlineKeyboardMarkup()
    # mede это med, просто чтобы в будущем коде скопировать релизацию time и не менять срезы
    btn_down = types.InlineKeyboardButton(text='↓', callback_data=f'mede_{time_type}down_{time_set}')
    btn_down5 = types.InlineKeyboardButton(text='-5', callback_data=f'mede_{time_type}dow5_{time_set}')
    btn_time = types.InlineKeyboardButton(text=f'{time_btn}',
                                          callback_data=f'mede_time')
    btn_up = types.InlineKeyboardButton(text='↑', callback_data=f'mede_{time_type}up_{time_set}')
    btn_up5 = types.InlineKeyboardButton(text='+5', callback_data=f'mede_{time_type}u5_{time_set}')
    btn_send = types.InlineKeyboardButton(text='Перейти к вводу времени начала', callback_data=f'mede_{time_type}send_{time_set}')
    inline_keyboard_time.add(btn_down, btn_time, btn_up)
    inline_keyboard_time.add(btn_down5, btn_up5)
    inline_keyboard_time.add(btn_send)
    return inline_keyboard_time


# Клавиатура с выбором времени администратором для водителей
def time_kb(time_set=time.time(), time_type='tm_', person=None, first=False):
    global change
    if first:
        change = person
    if time_type == 'dt_':
        time_btn = time.strftime("%d.%m", time.localtime(time_set))
    else:
        time_btn = time.strftime("%H:%M", time.localtime(time_set))
    inline_keyboard_time = types.InlineKeyboardMarkup()
    btn_down = types.InlineKeyboardButton(text='↓', callback_data=f'time_{time_type}down_{time_set}')
    btn_down5 = types.InlineKeyboardButton(text='-5', callback_data=f'time_{time_type}dow5_{time_set}')
    btn_time = types.InlineKeyboardButton(text=f'{time_btn}',
                                          callback_data=f'time_time')
    btn_up = types.InlineKeyboardButton(text='↑', callback_data=f'time_{time_type}up_{time_set}')
    btn_up5 = types.InlineKeyboardButton(text='+5', callback_data=f'time_{time_type}u5_{time_set}')
    btn_send = types.InlineKeyboardButton(text='Отправить', callback_data=f'time_{time_type}send_{time_set}')
    inline_keyboard_time.add(btn_down, btn_time, btn_up)
    inline_keyboard_time.add(btn_down5, btn_up5)
    inline_keyboard_time.add(btn_send)
    return inline_keyboard_time





# kb for create order
def keyboard_main():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    get_num = types.KeyboardButton('Получить номер')

    markup.add(get_num)
    return markup


# функция, для изменения расписания медиков и механников
def changer_med_meh(time_local, second = False):
    global change
    if second:
        cur_color = 'FFC000'
    else:
        cur_color = 'FFFF00'
    wb = load_workbook('ТабельМЕДИК-МЕХАНИК.xlsx')
    ws = wb.active
    row_local = ''
    column_leter = ''
    today = str(int(time.strftime('%d', time.localtime(time_local))))
    for cell in ws.iter_rows(min_col=1, max_col=1):
        if change == cell[0].value:
            row_local = cell[0].row
    for cell in ws.iter_cols(max_row=1):
        if cell[0].value == today:
            column_leter = cell[0].column_letter
    print(column_leter)
    print(row_local)
    ws[f'{column_leter}{row_local}'].fill = PatternFill(start_color=cur_color,
                                                        end_color=cur_color,
                                                        fill_type='solid')

    wb.save('ТабельМЕДИК-МЕХАНИК.xlsx')


# job for schedule: spam chat while user answer
def nerd(id):
    bot.send_message(id, "Вы не отправили фото путевого листа")


# reminder that the way list is ending
def hour_remind(id, first_name, last_name):
    wb = load_workbook(f'remind_report.xlsx')
    ws = wb.active
    local_count_row = ws.max_row + 1
    try:
        with open('log.txt', 'a') as f:
            f.write(
                f"{list({time.strftime('%H:%M', time.localtime())})[0]} Бот для {id}: часовое напоминание\n")
        bot.send_message(id,
                         "Медосмотр заканчивается через час. Получите номер путевого листа и заполните новый путевой лист.")
        id = int(id)
        ws[f'A{local_count_row}'] = first_name + '' + last_name
        ws[f'B{local_count_row}'].font = opx.styles.Font(name='Arial Cyr', charset=204, family=2.0, b=True,
                                                         color='FFC000', sz=8.0)
        ws[f'B{local_count_row}'] = 'Да'

    except:
        traceback.print_exc()
        ws[f'A{local_count_row}'] = first_name + last_name
        ws[f'B{local_count_row}'].font = opx.styles.Font(name='Arial Cyr', charset=204, family=2.0, b=True,
                                                         color='ff4e33', sz=8.0)
        ws[f'B{local_count_row}'] = 'Нет'
    with open(notif, 'r') as f:
        for line in f:
            if line.split() == None or line == '\n' or line.split()[0] == str(id):
                continue
            with open('notif_2.txt', 'a') as f2:
                f2.write(line)
    wb.save(f'remind_report.xlsx')
    os.remove(notif)
    os.rename('notif_2.txt', 'notif.txt')

    if os.path.exists('notif.txt'):
        pass
    else:
        with open("notif.txt", "w") as file:
            pass
    schedule.clear(f'{id}')
    schedule.every(30).minutes.do(hour_remind_2, id=id).tag(f'{id}')


def hour_remind_2(id):
    wb = load_workbook(f'remind_report.xlsx')
    ws = wb.active
    local_count_row = ws.max_row
    try:
        bot.send_message(id,
                         "Медоосмотр заканчивается через пол часа. Получите номер путевого листа и заполните новый путевой лист.")
        id = int(id)
        ws[f'C{local_count_row}'].font = opx.styles.Font(name='Arial Cyr', charset=204, family=2.0, b=True,
                                                         color='00db6a', sz=8.0)
        ws[f'C{local_count_row}'] = 'Да'
        with open('log.txt', 'a') as f:
            f.write(
                f"{list({time.strftime('%H:%M', time.localtime())})[0]} Бот для {id}: Получасовое напоминание\n")
        schedule.clear(f'{id}')
    except:
        traceback.print_exc()
        ws[f'C{local_count_row}'].font = opx.styles.Font(name='Arial Cyr', charset=204, family=2.0, b=True,
                                                         color='ff4e33', sz=8.0)
        ws[f'C{local_count_row}'] = 'Нет'
    wb.save(f'remind_report.xlsx')


def reminder(id, end_time, first_name, last_name):
    # end_time = end_time - 35900     \\Надстройка для теста напоминаний. Не вникай, доверься
    schedule.every().day.at(list({time.strftime("%H:%M", time.localtime(end_time))})[0]).do(hour_remind, id=id,
                                                                                            first_name=first_name,
                                                                                            last_name=last_name).tag(
        f'{id}')
    with open(notif, 'r') as f:
        for line in f:
            if line.split() == None or line == '\n' or line.split()[0] == str(id):
                continue
            with open('notif_2.txt', 'a') as f2:
                f2.write(line)
    with open('notif_2.txt', 'a') as f:
        f.write(f'{id} {end_time} {first_name} {last_name}\n')
    os.remove(notif)
    os.rename('notif_2.txt', 'notif.txt')


# get info from exel file about drivers
def excel_get(message, cmd):
    if os.path.exists(f'{driver_file_name}.xlsx'):
        try:
            wb = load_workbook(f'{driver_file_name}.xlsx')
            ws = wb.active
            number_column = ws['A']
            # bd data
            id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
            id_bd_cursor = id_bd_conn.cursor()
            id_bd_cursor.execute(f"SELECT gos_num FROM id_bd WHERE id = '{str(message.chat.id)}'")
            gos_num = id_bd_cursor.fetchone()
            id_bd_conn.commit()
            row_i = []
            for cell in number_column:
                if gos_num[0] == cell.value:
                    for row in ws.iter_rows(min_row=cell.row, max_col=11, max_row=cell.row):
                        for cell_i in row:
                            row_i.append(cell_i.value)
                    wb.save(f'{driver_file_name}.xlsx')
                    if cmd == 'name':
                        return row_i[5]
                    elif cmd == 'gos_num':
                        return row_i[0]
                    elif cmd == 'transport':
                        return row_i[1]
                    elif cmd == 'v/u':
                        return row_i[6]
            wb.save(f'{driver_file_name}.xlsx')
            id_bd_conn.close()
        except Exception as e:
            with open('error.txt', 'a') as f:
                f.write(str(e))
            traceback.print_exc()
            bot.send_message(message.chat.id, "Что-то пошло не так...")


# edit excel file
def excel_maker(message, start_time,med_time, med, meh):
    with open(path_to_n_cont, 'r') as f:
        n = int(f.read().strip())
    shapka = ['Номер путевого листа', 'Время начала', 'Время окончания', 'Мед. осмотр', 'ФИО сотрудника', 'Номер авто',
              'Транспорт',
              'Пробег', 'В/У','Медик','Механик']
    list_of_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','J','K']
    n_local = int(user_dict[message.chat.id]['local_n']) + 1
    if os.path.exists('report.xlsx'):
        try:
            wb_obj = load_workbook('report.xlsx')
            sheet_obj = wb_obj.active
            sheet_obj[f'A{n_local}'] = user_dict[message.chat.id]['global_n']

            sheet_obj[f'B{n_local}'] = list({time.strftime("%d.%m.%Y %H:%M", time.localtime(start_time))})[0]
            sheet_obj[f'C{n_local}'] = list({time.strftime("%d.%m.%Y %H:%M", time.localtime(start_time + 43199))})[0]
            sheet_obj[f'D{n_local}'] = list({time.strftime("%H:%M", time.localtime(med_time))})[0]
            sheet_obj[f'E{n_local}'] = excel_get(message, 'name')
            sheet_obj[f'F{n_local}'] = excel_get(message, 'gos_num')
            sheet_obj[f'G{n_local}'] = excel_get(message, 'transport')
            sheet_obj[f'H{n_local}'] = user_dict[message.chat.id]['mileage']
            sheet_obj[f'I{n_local}'] = excel_get(message, 'v/u')
            sheet_obj[f'J{n_local}'] = med
            sheet_obj[f'K{n_local}'] = meh
            wb_obj.save('report.xlsx')
        except Exception as e:
            with open('error.txt', 'a') as f:
                f.write(f'{str(e)}\n')
        return start_time
    else:
        # create excel file if it don't exist
        wb = Workbook()
        sheet = wb.active
        border = Border(left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))
        for i in range(0, len(shapka)):
            sheet[f'{list_of_letters[i]}1'] = shapka[i]
            sheet[f'{list_of_letters[i]}1'].fill = PatternFill(start_color='7fc7ff',
                                                               end_color='7fc7ff',
                                                               fill_type='solid')
            sheet[f'{list_of_letters[i]}1'].border = border
            sheet.column_dimensions[f'{list_of_letters[i]}'].width = len(shapka[i]) + 5
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['I'].width = 20
        sheet.row_dimensions[1].height = 20
        wb.save('report.xlsx')
        return excel_maker(message, start_time, med_time, med, meh)


# button click handler
@bot.callback_query_handler(func=lambda call: True)
def callback_worker(call):
    global time_change, change
    with open('log.txt', 'a') as f:
        f.write(
            f"{list({time.strftime('%H:%M', time.localtime())})[0]}, {call.message.chat.id},{call.message.from_user.first_name} {call.message.from_user.last_name}: нажал кнопку {call.data}\n")
    # bd data
    id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
    id_bd_cursor = id_bd_conn.cursor()
    if call.data[:3] == "reg":
        if call.data[4:] == 'del':
            id_bd_cursor.execute(f"DELETE FROM id_bd WHERE id ='{(str(call.message.chat.id))}'")
            id_bd_conn.commit()
            bot.send_message(call.message.chat.id, "Данные удалены", reply_markup=keyboard_main())
        elif call.data[4:] == 'change':
            id_bd_cursor.execute(f"DELETE FROM id_bd WHERE id ='{(str(call.message.chat.id))}'")
            id_bd_conn.commit()
            bot.send_message(call.message.chat.id, "Нажмите на кнопку отправить номер", reply_markup=phone_btn())
            bot.register_next_step_handler(call.message, get_num_for_reg)
    elif call.data[:7] == 'worker_':

        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Выберите дату',
                              reply_markup=time_kb(first=True, person=call.data[7:], time_type='dt_',
                                                   time_set=time.time()))
    elif call.data[:4] == 'chw_':
        change = call.data[4:]
        changer_med_meh(float(time_change))
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Готово',
                              reply_markup=None)
    elif call.data[:5] == 'mede_':
        difference = 60
        difference5 = 300
        if call.data[8:13] == 'down_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_med_kb(time_set=(float(call.data[13:])) - difference,
                                                           time_type=call.data[5:8]))
        elif call.data[8:13] == 'dow5_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_med_kb(time_set=(float(call.data[13:])) - difference5,
                                                           time_type=call.data[5:8]))
        elif call.data[8:11] == 'up_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_med_kb(time_set=(float(call.data[11:])) + difference))
        elif call.data[8:11] == 'u5_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_med_kb(time_set=(float(call.data[11:])) + difference5))

        elif call.data[8:13] == 'send_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=f'{call.message.text}\n'
                                       f''
                                       f'{time.strftime("%H:%M", time.localtime(float(call.data[13:])))}    \n  <b>Мед осмотр </b>',
                                  reply_markup=None, parse_mode='HTML')
            message_dict[call.message.message_id]['med_time'] = float(call.data[13:])
            bot.send_message(chat_id=call.message.chat.id,
                             text=f'Введите время начала',
                             reply_markup=time_kb(time_set=time.time(), time_type='tm_'), parse_mode='HTML')
    elif call.data[:5] == 'time_':
        if call.data[5:8] == 'tm_':
            difference = 60
            difference5 = 300
        else:
            difference = 86400
            difference5 = 432000
        if call.data[8:13] == 'down_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_kb(time_set=(float(call.data[13:])) - difference,
                                                       time_type=call.data[5:8]))
        elif call.data[8:13] == 'dow5_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_kb(time_set=(float(call.data[13:])) - difference5,
                                                       time_type=call.data[5:8]))
        elif call.data[8:11] == 'up_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_kb(time_set=(float(call.data[11:])) + difference,
                                                       time_type=call.data[5:8]))
        elif call.data[8:11] == 'u5_':
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text=call.message.text,
                                  reply_markup=time_kb(time_set=(float(call.data[11:])) + difference5,
                                                       time_type=call.data[5:8]))
        elif call.data[8:13] == 'send_':
            if call.data[5:8] == 'tm_':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text=f'{call.message.text}\n'
                                           f''
                                           f'{time.strftime("%H:%M", time.localtime(float(call.data[13:])))}    \n  <b>Отправленно</b>',
                                      reply_markup=None, parse_mode='HTML')
                get_time(float(call.data[13:]),
                         bot.send_message(chat_id=message_dict[call.message.message_id - 1]['chat'],
                                          text='Время отправления:'), float(message_dict[call.message.message_id-1]['med_time']))
            else:
                changer_med_meh(float(call.data[13:]), second=True)
                time_change = call.data[13:]
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text=f'{call.message.text}\n'
                                           f''
                                           f'{time.strftime("%H:%M", time.localtime(float(call.data[13:])))}    \n  <b>Изменено</b>\n'
                                           f'Выберете работника на замену',
                                      reply_markup=worker_change_kb(), parse_mode='HTML')


    elif call.data == "cancel":
        bot.send_message(call.message.chat.id,
                         'Чтобы получить номер путевого листа нажмите на кнопку снизу, либо введите "Получить номер"',
                         reply_markup=keyboard_main())
        id_bd_conn.close()


def get_sur_med():
    wb = load_workbook('ТабельМЕДИК-МЕХАНИК.xlsx')
    ws = wb.active
    # берем сегодняшнее число
    today = str(int(time.strftime('%d', time.localtime())))
    now_hour = int(time.strftime('%H', time.localtime()))
    now_min = int(time.strftime('%M', time.localtime()))
    for cell in ws.iter_cols(max_row=1):
        # pprint(dir(cell[0]))
        # break
        if cell[0].value == today:
            column_leter = cell[0].column_letter
    if 8 < now_hour < 20:
        # какая сейчас смена
        shift = 'день'
    else:
        shift = 'ночь'
    row_mas = []
    for cell in ws.iter_rows(min_col=3, max_col=3):
        if cell[0].value == shift or cell[0].value == 'сутки':
            row_mas.append(cell[0].row)
    meh_mas = []
    med_mas = []
    for cell in ws.iter_rows(min_col=2, max_col=2):
        if cell[0].value == 'медик':
            med_mas.append(cell[0].row)
        if cell[0].value == 'механик':
            meh_mas.append(cell[0].row)
    med_mas = set(med_mas) & set(row_mas)
    meh_mas = set(meh_mas) & set(row_mas)
    worker_dict = {'Механик': '',
                   'Медик': ''}
    for el in med_mas:
        if ws[f'{column_leter}{el}'].fill.fgColor.value[2:] == 'FFFF00':
            worker_dict['Медик'] = f"{worker_dict['Медик']} {ws[f'A{el}'].value}"
    for el in meh_mas:
        if ws[f'{column_leter}{el}'].fill.fgColor.value[2:] == 'FFFF00':
            worker_dict['Механик'] = f"{worker_dict['Механик']} {ws[f'A{el}'].value}"
    wb.save('ТабельМЕДИК-МЕХАНИК.xlsx')
    return worker_dict


# listening server
@bot.message_handler(content_types=['text', 'photo', 'document'])
# main cmds
def start(message):
    # bd data
    with open('log.txt', 'a') as f:
        f.write(
            f"{list({time.strftime('%H:%M', time.localtime())})[0]} {message.from_user.id} {message.from_user.first_name} {message.from_user.last_name}: {message.text}\n")
    id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
    id_bd_cursor = id_bd_conn.cursor()
    if message.from_user.id == s_k_id or message.from_user.id == mng_id:
        if message.text == 'Обновить список водителей':
            bot.send_message(message.from_user.id, "Пришлите excel файл водителей", reply_markup=None)
            bot.register_next_step_handler(message, refresh_file)
            id_bd_conn.close()
            return
        elif message.text == 'Обновить список работников':
            bot.send_message(message.from_user.id, "Пришлите excel файл медиков и механиков", reply_markup=None)
            bot.register_next_step_handler(message, refresh_med_meh)
            id_bd_conn.close()
            return
        elif message.text == 'Замена сотрудника':
            bot.send_message(message.from_user.id, "Выберите сотрудника", reply_markup=worker_kb())
            return

        elif message.text == 'Рассылка':
            bot.send_message(message.from_user.id, "Напишите список номеров", reply_markup=None)
            bot.register_next_step_handler(message, spam)
            id_bd_conn.close()
            return
        elif message.text == 'Вернуться':
            bot.send_message(message.from_user.id,
                             """ 
                             Возвращаюсь
                             """, reply_markup=keyboard_main())
            return
        elif message.text == '/admin':
            bot.send_message(message.from_user.id, "Вы не зарегестрированы", reply_markup=kb_cmd_admin())
            return
    for row in id_bd_cursor.execute("SELECT id FROM id_bd"):
        if str(message.from_user.id) == row[0]:
            if message.text == 'Получить номер':
                user_dict[message.chat.id] = {'send_photo': None, 'local_n': None, 'global_n': None, 'time_start': None,
                                              'mileage': None}
                bot.send_message(message.from_user.id, "Введите ваш пробег", reply_markup=None)
                bot.register_next_step_handler(message, get_mileage)
            elif message.text == '/reg':
                check_on_reg(message)
            else:
                # bot.send_message(message.from_user.id, "Не понял Вас", reply_markup=phone_btn())
                # todo вернуть это
                bot.send_message(message.from_user.id, "Не понял Вас", reply_markup=keyboard_main())
                return

            id_bd_conn.close()
            break
    else:
        if message.text == '/start':
            bot.send_message(message.from_user.id,
                             """ 
                             Работаю
                             """, reply_markup=keyboard_main())
        elif message.text == '/reg':
            check_on_reg(message)
        else:
            bot.send_message(message.from_user.id, "Вы не зарегестрированы")


def spam(message):
    miss_driver = list(message.text.split())
    bot.send_message(message.chat.id, "Какое сообщение отправить?", reply_markup=None)
    bot.register_next_step_handler(message, spam2, miss_driver)


def spam2(message, drivers):
    id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
    id_bd_cursor = id_bd_conn.cursor()
    id_list = []
    palm = id_bd_cursor.execute("SELECT * FROM id_bd").fetchall()
    id_bd_conn.commit()
    id_bd_conn.close()
    driver_dict = {}
    wrong_num = ''
    for el in palm:
        driver_dict[el[1]] = el[0]
    for num in drivers:
        try:
            id_list.append(driver_dict[num])
        except KeyError:
            wrong_num += num + '\n'
    if wrong_num != '':
        bot.send_message(message.chat.id, f'Этих номеров нет в базе данных\n{wrong_num}')
    msg = message.text
    t2 = threading.Thread(target=spam_theard, name='schedule', args=(msg, id_list))
    t2.start()
    del palm
    del driver_dict
    del wrong_num


def spam_theard(msg, drivers):
    time.sleep(1)
    for each_id in drivers:
        bot.send_message(each_id, msg)
        time.sleep(1)


def refresh_file(message):
    if message.document != None:
        if message.document.file_name.endswith('xlsx'):
            file_info_excel = bot.get_file(message.document.file_id)
            downloaded_file_excel = bot.download_file(file_info_excel.file_path)
            with open(f'driver_list.xlsx', 'wb') as f:
                f.write(downloaded_file_excel)
            bot.send_message(message.chat.id, 'файл принят')
        else:
            bot.send_message(message.chat.id, 'Не правильный тип файла')
    else:
        bot.send_message(message.chat.id, 'нужно прислать файл')


def refresh_med_meh(message):
    if message.document != None:
        if message.document.file_name.endswith('xlsx'):
            file_info_excel = bot.get_file(message.document.file_id)
            downloaded_file_excel = bot.download_file(file_info_excel.file_path)
            with open(f'ТабельМЕДИК-МЕХАНИК.xlsx', 'wb') as f:
                f.write(downloaded_file_excel)
            bot.send_message(message.chat.id, 'файл принят')
        else:
            bot.send_message(message.chat.id, 'Не правильный тип файла')
    else:
        bot.send_message(message.chat.id, 'нужно прислать файл')


def check_on_reg(message):
    # bd data
    id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
    id_bd_cursor = id_bd_conn.cursor()
    for row in id_bd_cursor.execute("SELECT id FROM id_bd"):
        if str(message.chat.id) == row[0]:
            bot.send_message(message.chat.id, "Вы хотите обновить или удалить свои данные?",
                             reply_markup=reg_keyboard())
            id_bd_conn.commit()
            return
    id_bd_conn.commit()
    id_bd_conn.close()
    bot.send_message(message.chat.id, "Нажмите на кнопку отправить номер", reply_markup=phone_btn())
    bot.register_next_step_handler(message, get_num_for_reg)


def get_mileage(message):
    with open('log.txt', 'a') as f:
        f.write(
            f"{list({time.strftime('%H:%M', time.localtime())})[0]} {message.from_user.id} {message.from_user.first_name}: {message.from_user.last_name} {message.text}\n")
    if message.text != 'Получить номер' and message.text != None:
        user_dict[message.chat.id]['mileage'] = message.text
    else:
        bot.send_message(message.chat.id, "Введите ваш пробег", reply_markup=None)
        bot.register_next_step_handler(message, get_mileage)
        return
    bot.send_message(message.chat.id, 'Подождите, диспетчер выдаст время')
    with open(path_to_n_cont, 'r') as f:
        n = f.read()
        n = int(n.strip())
    n = n + 1
    user_dict[message.chat.id]['local_n'] = int(n)

    with open(path_to_n_cont, 'w') as f:
        f.write(str(n))

    with open(path_n_global, 'r') as f:
        ng = f.read()
        ng = int(ng.strip())
    ng = ng + 1
    user_dict[message.chat.id]['global_n'] = int(ng)
    with open(path_n_global, 'w') as f:
        f.write(str(ng))
    id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
    id_bd_cursor = id_bd_conn.cursor()
    id_bd_cursor.execute(f"SELECT gos_num FROM id_bd WHERE id = {(str(message.chat.id))}")
    message_id_tuk = bot.send_message(mng_id,
                                  f'Номер авто: {id_bd_cursor.fetchone()[0]} \n'
                                  f'Номер путевого листа: {user_dict[message.chat.id]["global_n"]}\n'
                                  '             Введите время МЕД ОСМОТРА',
                                  parse_mode='HTML', reply_markup=time_med_kb(time_set=time.time())).message_id
    message_dict[message_id_tuk] = {}
    message_dict[message_id_tuk]['chat'] = message.chat.id
    id_bd_conn.commit()
    id_bd_conn.close()


def get_time(time_from_mng, message, med_time):
    id_chat = message.chat.id
    worker_dict = get_sur_med()
    start_time = excel_maker(message, time_from_mng, med_time,worker_dict['Медик'], worker_dict['Механик'])
    user_dict[id_chat]['time_start'] = start_time
    bot.send_message(id_chat,
                     f"ВНИМАНИЕ, Сейчас\n"
                     f"Медик - {worker_dict['Медик']}\n"
                     f"Механик - {worker_dict['Механик']}\n\n"
                     f"Номер: \n    "
                     f"<b>{str(user_dict[id_chat]['global_n']).center(48, ' ')}</b>\n\n"
                     f"Время начала: <b>{list({time.strftime('%d.%m.%Y %H:%M', time.localtime(start_time))})[0]}</b>\n\n"
                     f"Окончание: <b>{list({time.strftime('%d.%m.%Y %H:%M', time.localtime(start_time + 43199))})[0]}</b>\n\n"
                     f"Начало медосмотра: <b>{list({time.strftime('%d.%m.%Y %H:%M', time.localtime(med_time))})[0]}</b>\n\n"
                     f"Контроль тех сост: <b>{list({time.strftime('%d.%m.%Y %H:%M', time.localtime(start_time))})[0]}</b>\n"
                     f"Выезд с парковки: <b>{list({time.strftime('%d.%m.%Y %H:%M', time.localtime(start_time))})[0]}</b>\n\n\n"
                     f"<b>{'Пришлите фото путевого листа'.center(48, ' ')}</b>\n",
                     parse_mode='HTML')

    # remember that user dont send photo
    schedule.every(30).minutes.do(nerd, id=id_chat).tag(f'{id_chat}')
    # bot.send_message(message.from_user.id, "Пришлите фото путевого листа", reply_markup=None)
    bot.register_next_step_handler(message, get_photo, start_time)


def get_photo(message, start_time):
    if message.photo != None:
        # photo
        with open('log.txt', 'a') as f:
            try:
                f.write(
                    f"{list({time.strftime('%H:%M', time.localtime())})[0]} {message.from_user.id} {message.from_user.first_name} {message.from_user.last_name}: \"фото {int(user_dict[message.chat.id]['global_n'])}\"\n")
            except Exception as e:
                traceback.print_exc()
                with open('error.txt', 'a') as f:
                    f.write(str(e))
        try:
            schedule.clear(f'{message.chat.id}')
            file_info = bot.get_file(message.photo[-1].file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            src = f'photo/{date.today()}___{int(user_dict[message.chat.id]["global_n"])}.jpg'
            user_dict[message.chat.id]['send_photo'] = src
            with open(src, 'wb') as new_file:
                new_file.write(downloaded_file)

        except Exception as es:
            bot.send_message(message.chat.id,
                             'Что-то пошло не так, повторите попытку')
            with open('error.txt', 'a') as f:
                f.write(str(es))

        try:
            if message.from_user.first_name != None:
                name = message.from_user.first_name
            else:
                name = '_'
            if message.from_user.last_name != None:
                surname = message.from_user.last_name
            else:
                surname = '_'
            reminder(message.chat.id, start_time + 39600, name, surname)
        except Exception as es:
            with open('error.txt', 'a') as f:
                f.write(str(es))
            traceback.print_exc()
        id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
        id_bd_cursor = id_bd_conn.cursor()
        id_bd_cursor.execute(f"SELECT gos_num FROM id_bd WHERE id = {(str(message.chat.id))}")
        bot.send_photo(chat_id=mng_id,
                       photo=open(user_dict[message.chat.id]['send_photo'], 'rb'),
                       caption=f'Номер авто: <b>{id_bd_cursor.fetchone()[0]}</b> \n'
                               f'Номер путевого листа: <b>{user_dict[message.chat.id]["global_n"]}</b>\n'
                               f'Начало действия путевого листа: <b>{list({time.strftime("%H:%M", time.localtime(start_time))})[0]}</b>\n',
                       parse_mode='HTML')
        id_bd_conn.commit()
        id_bd_conn.close()

        bot.send_message(message.chat.id,
                         'Помните, езда без путевого листа запрещена,<b> за отсутвие путевого листа штраф 2000 р.</b>\n Хорошего пути!',
                         parse_mode='HTML')
        del user_dict[message.chat.id]
    else:
        bot.send_message(message.chat.id, 'Пришлите фотографию')
        bot.register_next_step_handler(message, get_photo)


def get_num_for_reg(message):
    if message.contact.phone_number == None:
        bot.send_message(message.chat.id, "Нажмите на кнопку отправить номер", reply_markup=phone_btn())
        bot.register_next_step_handler(message, get_num_for_reg)
    phone_num = message.contact.phone_number
    # check whether the auto is in organization file excel
    if os.path.exists(f'{driver_file_name}.xlsx'):
        try:
            id_bd_conn = sqlite3.connect('drivers_id.sqlite', timeout=1)
            id_bd_cursor = id_bd_conn.cursor()
            wb = load_workbook(f'{driver_file_name}.xlsx')
            ws = wb.active
            #ищем номер телефона по xls файлу в столбце E
            for cell in ws.iter_rows(min_col=5, max_col=5):
                phone_from_xls = ''
                if cell[0].value == None:
                    continue
                for el in re.findall(r'\d+', str(cell[0].value)):
                    phone_from_xls = phone_from_xls + el
                    if phone_from_xls == phone_num:
                        try:
                            id_bd_cursor.execute(
                                f"INSERT INTO id_bd VALUES ('{str(message.chat.id)}','{str(ws[f'A{cell[0].row}'].value)}', '{str(phone_num)}')")
                            id_bd_conn.commit()
                            bot.send_message(message.chat.id,
                                             f'Вы зарегестированны.\n'
                                             f'Ваше имя: {str(ws[f"F{cell[0].row}"].value)}\n'
                                             f'Номер машины: {str(ws[f"A{cell[0].row}"].value)}\n'
                                             f' \n Чтобы получить номер путевого листа нажмите на кнопку снизу, либо введите "Получить номер"',
                                             reply_markup=keyboard_main())
                            with open('log.txt', 'a') as f:
                                f.write(
                                    f"{list({time.strftime('%H:%M', time.localtime())})[0]}  Бот для {message.from_user.id}: {ws[f'F{cell[0].row}'].value}, \n указанный номер верен? \n <b>{cell[0].value}</b>")
                            wb.save(f'{driver_file_name}.xlsx')
                            id_bd_conn.close()
                            return
                        except Exception as es:
                            with open('error.txt', 'a') as f:
                                f.write(str(es))
                            traceback.print_exc()
            wb.save(f'{driver_file_name}.xlsx')
            id_bd_conn.close()
            bot.send_message(message.chat.id,
                             'Ваш номер не найден в списке', reply_markup=keyboard_main())
        except Exception as es:
            with open('error.txt', 'a') as f:
                f.write(str(es))
            traceback.print_exc()


def schedule_theard():
    while True:
        time.sleep(5)
        schedule.run_pending()


if __name__ == '__main__':
    t = threading.Thread(target=schedule_theard, name='schedule')
    t.start()
    try:
        with open(notif, 'r') as f:
            for line in f:
                if 'start' in line or line.split() == None or line == '\n':
                    continue
                schedule.every().day.at(list({time.strftime("%H:%M", time.localtime(float(line.split()[1])))})[0]).do(
                    hour_remind,
                    id=line.split()[0], first_name=line.split()[2], last_name=line.split()[3]).tag(f'{line.split()[0]}')
    except:
        traceback.print_exc()
    while True:
        try:
            bot.polling(none_stop=True, interval=1)
        except Exception as e:
            with open('error.txt', 'a') as f:
                f.write(str(e))
            traceback.print_exc()
            time.sleep(5)
