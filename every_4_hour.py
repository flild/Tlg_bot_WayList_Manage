import schedule
import time
import telebot

from config import token_b, s_k_id, mng_id, path_dir
from openpyxl import load_workbook

#os.chdir(path_dir)
def job():
    try:
        miss_driver = ''
        count=0
        wb_rep = load_workbook('report.xlsx')
        ws_rep = wb_rep.active
        number_column_rep = ws_rep['F']
        wb_dr = load_workbook('driver_list.xlsx')
        ws_dr = wb_dr.active
        number_column_dr = ws_dr['A']
        count_dr = len(number_column_dr)-2
        for driver in number_column_dr:
            for rep in number_column_rep:
                if driver.value == rep.value or driver.value=='end' or driver.value=='Гос номер':
                    break
            if driver.value == rep.value or driver.value == 'end' or driver.value == 'Гос номер':
                continue
            count +=1
            miss_driver = miss_driver + f'{driver.value}\n'
        bot = telebot.TeleBot(token_b)
        bot.send_message(s_k_id, f'Не вышло водителей {count}/{count_dr}:\n{miss_driver}')
        bot.send_message(mng_id, f'Не вышло водителей {count}/{count_dr}:\n{miss_driver}')
        del miss_driver
        wb_rep.save(f'report.xlsx')
        wb_dr.save(f'driver_list.xlsx')
    except:
        del miss_driver



schedule.every().day.at("23:58").do(job)
schedule.every().day.at("04:00").do(job)
schedule.every().day.at("08:00").do(job)
schedule.every().day.at("12:00").do(job)
schedule.every().day.at("16:00").do(job)
schedule.every().day.at("20:00").do(job)


while True:
    schedule.run_pending()
    # txt = input()
    # if txt == 'send':
    #     job()
    time.sleep(60)